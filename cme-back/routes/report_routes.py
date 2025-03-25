# routes/report_routes.py
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
from io import BytesIO
from fpdf import FPDF
from database import get_db
from models.process import Process
from models.failure import Failure
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from middlewares.auth import verify_jwt
import pandas as pd
import os

router = APIRouter()

@router.get("/pdf", response_class=StreamingResponse, dependencies=[Depends(verify_jwt)])
def generate_pdf_report(db: Session = Depends(get_db)):
    processos = db.query(Process).all()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    def write_stage(title, data, obs):
        if data:
            formatted = data.strftime("%d/%m/%Y %H:%M")
            text = f"{title}: {formatted} - {obs or 'Sem observações'}"
        else:
            text = f"{title}: Não realizada"
        pdf.set_x(10)
        pdf.multi_cell(0, 8, txt=text)

    for processo in processos:
        pdf.set_font("Arial", "B", 12)
        pdf.set_x(10)
        pdf.cell(0, 10, f"Processo ID: {processo.id}", ln=True)

        pdf.set_font("Arial", "", 12)
        pdf.set_x(10)
        pdf.cell(0, 8, f"Material ID: {processo.material_id}", ln=True)

        write_stage("Recebimento", processo.data_recebimento, processo.observacao_recebimento)
        write_stage("Lavagem", processo.data_lavagem, processo.observacao_lavagem)
        write_stage("Esterilização", processo.data_esterilizacao, processo.observacao_esterilizacao)
        write_stage("Distribuição", processo.data_distribuicao, processo.observacao_distribuicao)

        failures = db.query(Failure).filter(Failure.process_id == processo.id).all()
        if failures:
            pdf.set_x(10)
            pdf.cell(0, 10, "Falhas:", ln=True)
            for falha in failures:
                critica = "Sim" if falha.critical else "Não"
                texto = f"• {falha.etapa.capitalize()} - {'Crítica' if falha.critical else 'Comum'}: {falha.descricao} ({falha.data.strftime('%d/%m/%Y %H:%M')})"
                pdf.set_x(10)
                pdf.multi_cell(0, 8, txt=texto)

        pdf.ln(5)

    buffer = BytesIO()
    pdf.output(buffer)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": "inline; filename=relatorio_processos.pdf"
    })

@router.get("/xlsx", response_class=FileResponse, dependencies=[Depends(verify_jwt)])
def generate_xlsx_report(db: Session = Depends(get_db)):
    processos = db.query(Process).options(
        joinedload(Process.material),
        joinedload(Process.failures),
        joinedload(Process.usuario_recebimento),
        joinedload(Process.usuario_lavagem),
        joinedload(Process.usuario_esterilizacao),
        joinedload(Process.usuario_distribuicao)
    ).all()

    data = []
    for processo in processos:
        row = {
            "ID Processo": processo.id,
            "Material": f"{processo.material.nome} (ID: {processo.material.id})",
            "Serial": processo.material.serial,
            "Recebimento": format_stage(processo.data_recebimento, processo.observacao_recebimento, processo.usuario_recebimento),
            "Lavagem": format_stage(processo.data_lavagem, processo.observacao_lavagem, processo.usuario_lavagem),
            "Esterilização": format_stage(processo.data_esterilizacao, processo.observacao_esterilizacao, processo.usuario_esterilizacao),
            "Distribuição": format_stage(processo.data_distribuicao, processo.observacao_distribuicao, processo.usuario_distribuicao),
            "Falhas": format_failures(processo.failures),
            "Interrompido": "Sim" if any(f.critical for f in processo.failures) else "Não"
        }
        data.append(row)

    df = pd.DataFrame(data)

    filename = f"relatorio_processos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join("relatorios", filename)
    os.makedirs("relatorios", exist_ok=True)
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=8)
        cell.alignment = Alignment(wrap_text=True)

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(path)

    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)

def format_stage(data, obs, user):
    if not data:
        return "-"
    user_info = f"{user.name} (ID: {user.id})" if user else "Usuário desconhecido"
    obs_text = f" - {obs}" if obs else ""
    return f"{data.strftime('%d/%m/%Y %H:%M')} - {user_info}{obs_text}"

def format_failures(failures):
    if not failures:
        return "-"
    lines = []
    for f in failures:
        tipo = "Crítica" if f.critical else "Comum"
        data = f.data.strftime('%d/%m/%Y %H:%M') if f.data else "Data desconhecida"
        lines.append(f"{f.etapa.capitalize()} - {tipo}: {f.descricao} ({data})")
    return "\n".join(lines)
